import html
import io
import re
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st


st.set_page_config(
    page_title="Buku Klapper Dapodik",
    page_icon="📘",
    layout="wide",
)

st.markdown(
    """
    <style>
    .watermark {
        position: fixed;
        bottom: 15px;
        right: 20px;
        font-size: 13px;
        color: rgba(0, 0, 0, 0.35);
        z-index: 9999;
        font-weight: 500;
    }
    </style>

    <div class="watermark">
        Tun Pasek Sarwiko Dipranoto<br>
        Universitas Udayana
    </div>
    """,
    unsafe_allow_html=True
)

# =========================
# Konfigurasi kolom aplikasi
# =========================

def jumlah_kelas(jenjang: str) -> int:
    return 6 if jenjang == "SD" else 3


def class_columns(jenjang: str) -> list[str]:
    return [f"Kelas {i}" for i in range(1, jumlah_kelas(jenjang) + 1)]


def all_columns(jenjang: str) -> list[str]:
    return [
        "No Urut",
        "No Induk",
        "NISN",
        "Nama siswa",
        "L/P",
        "Tempat Tanggal Lahir",
        "Nama Orang Tua",
        *class_columns(jenjang),
        "Tanggal Keluar Sekolah",
        "Catatan",
    ]


ALIASES = {
    "no": "No Urut",
    "no.": "No Urut",
    "nomor": "No Urut",
    "nomor urut": "No Urut",
    "no urut": "No Urut",
    "urut": "No Urut",
    "induk": "No Induk",
    "no induk": "No Induk",
    "nomor induk": "No Induk",
    "nis": "No Induk",
    "nipd": "No Induk",
    "nisn": "NISN",
    "nama": "Nama siswa",
    "nama siswa": "Nama siswa",
    "nama peserta didik": "Nama siswa",
    "peserta didik": "Nama siswa",
    "l/p": "L/P",
    "lp": "L/P",
    "jk": "L/P",
    "jenis kelamin": "L/P",
    "tempat tanggal lahir": "Tempat Tanggal Lahir",
    "tempat tgl lahir": "Tempat Tanggal Lahir",
    "ttl": "Tempat Tanggal Lahir",
    "tempat lahir": "Tempat Tanggal Lahir",
    "tanggal lahir": "Tempat Tanggal Lahir",
    "tgl lahir": "Tempat Tanggal Lahir",
    "nama orang tua": "Nama Orang Tua",
    "orang tua": "Nama Orang Tua",
    "nama ortu": "Nama Orang Tua",
    "ayah": "Nama Orang Tua",
    "ibu": "Nama Orang Tua",
    "tanggal keluar": "Tanggal Keluar Sekolah",
    "tanggal keluar sekolah": "Tanggal Keluar Sekolah",
    "tgl keluar": "Tanggal Keluar Sekolah",
    "tanggal lulus": "Tanggal Keluar Sekolah",
    "tgl lulus": "Tanggal Keluar Sekolah",
    "keluar sekolah": "Tanggal Keluar Sekolah",
    "keterangan": "Catatan",
    "catatan": "Catatan",
}


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split())


def normalize_column_name(name) -> str:
    text = clean_text(name)
    key = text.lower().replace("_", " ").replace("-", " ")
    key = " ".join(key.split())

    # Kolom tanggal naik/masuk kelas bisa ditulis dengan banyak variasi.
    for i in range(1, 7):
        patterns = {
            str(i),
            f"kelas {i}",
            f"kls {i}",
            f"tanggal naik kelas {i}",
            f"tgl naik kelas {i}",
            f"tanggal naik/masuk kelas {i}",
            f"tgl naik/masuk kelas {i}",
            f"naik kelas {i}",
            f"masuk kelas {i}",
        }
        if key in patterns:
            return f"Kelas {i}"

    return ALIASES.get(key, text)


def blank_dataframe(jenjang: str, n_rows: int = 10) -> pd.DataFrame:
    return pd.DataFrame([{col: "" for col in all_columns(jenjang)} for _ in range(n_rows)])


def remove_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    temp = df.copy().replace("", pd.NA)
    temp = temp.dropna(how="all")

    # Hilangkan baris footer/template kosong yang kadang hanya berisi teks di kolom No Urut,
    # misalnya watermark sumber template. Baris dianggap data jika minimal punya
    # No Induk/NISN/Nama/TTL/Orang Tua/Catatan atau salah satu kolom kelas.
    key_cols = [
        c for c in temp.columns
        if c != "No Urut" and (c in ["No Induk", "NISN", "Nama siswa", "Tempat Tanggal Lahir", "Nama Orang Tua", "Tanggal Keluar Sekolah", "Catatan"] or str(c).startswith("Kelas "))
    ]
    if key_cols:
        temp = temp.dropna(subset=key_cols, how="all")
    return temp.fillna("")


def normalize_dataframe(df: pd.DataFrame, jenjang: str) -> pd.DataFrame:
    if df is None or df.empty:
        return blank_dataframe(jenjang)

    df = df.copy()
    df.columns = [normalize_column_name(c) for c in df.columns]

    # Gabungkan kolom Tempat Lahir + Tanggal Lahir bila pengguna mengunggah data mentah Dapodik.
    if "Tempat Tanggal Lahir" not in df.columns:
        tempat_cols = [c for c in df.columns if str(c).lower() in ["tempat lahir", "tempat"]]
        tgl_cols = [c for c in df.columns if str(c).lower() in ["tanggal lahir", "tgl lahir"]]
        if tempat_cols or tgl_cols:
            tempat = df[tempat_cols[0]].astype(str).replace("nan", "") if tempat_cols else ""
            tanggal = df[tgl_cols[0]].apply(format_excel_date) if tgl_cols else ""
            df["Tempat Tanggal Lahir"] = [gabung_dua(a, b, pemisah=", ") for a, b in zip(tempat, tanggal)]

    # Gabungkan Nama Ayah + Nama Ibu bila tersedia.
    if "Nama Orang Tua" not in df.columns:
        ayah = None
        ibu = None
        for c in df.columns:
            low = str(c).lower()
            if low in ["nama ayah", "ayah"]:
                ayah = c
            if low in ["nama ibu", "ibu"]:
                ibu = c
        if ayah or ibu:
            ayah_val = df[ayah].astype(str).replace("nan", "") if ayah else ""
            ibu_val = df[ibu].astype(str).replace("nan", "") if ibu else ""
            df["Nama Orang Tua"] = [gabung_dua(a, b, pemisah=" / ") for a, b in zip(ayah_val, ibu_val)]

    # Jika No Urut tidak ada, buat otomatis.
    if "No Urut" not in df.columns:
        df.insert(0, "No Urut", range(1, len(df) + 1))

    for col in all_columns(jenjang):
        if col not in df.columns:
            df[col] = ""

    df = df[all_columns(jenjang)]
    df = df.fillna("")
    return df


def gabung_dua(a, b, pemisah=", ") -> str:
    a = clean_text(a)
    b = clean_text(b)
    if a and b:
        return f"{a}{pemisah}{b}"
    return a or b


def format_excel_date(value) -> str:
    """Ubah nilai menjadi teks aman untuk tampilan/export.

    Catatan penting: angka seperti No Urut = 1 tidak boleh dianggap tanggal
    oleh pandas, karena bisa berubah menjadi 01/01/1970. Karena itu angka
    dikembalikan sebagai angka/teks biasa.
    """
    if pd.isna(value) or value == "":
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            return str(int(value))
        return str(value)

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""

    # Parse hanya jika teks tampak seperti tanggal, bukan semua teks/angka.
    looks_like_date = bool(re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text))
    if looks_like_date:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
        if not pd.isna(parsed):
            return parsed.strftime("%d/%m/%Y")
    return text


def sort_dataframe(df: pd.DataFrame, sort_by: str) -> pd.DataFrame:
    if sort_by not in df.columns or df.empty:
        return df

    df = df.copy()
    if sort_by in ["No Urut", "No Induk", "NISN"]:
        helper = df[sort_by].astype(str).str.extract(r"(\d+)", expand=False)
        helper = pd.to_numeric(helper, errors="coerce")
    else:
        helper = df[sort_by].astype(str).str.lower()

    df["_helper"] = helper
    df = df.sort_values(["_helper", sort_by], ascending=True, na_position="last", kind="mergesort")
    return df.drop(columns=["_helper"]).reset_index(drop=True)


# =========================
# Pembacaan file upload
# =========================

def read_excel_upload(uploaded_file, sheet_name: str, jenjang: str) -> pd.DataFrame:
    """Membaca dua kemungkinan format:
    1. Format Buku Klapper cetak dengan header 2 baris seperti contoh pengguna.
    2. Format data mentah/table biasa dengan baris pertama sebagai nama kolom.
    """
    raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, dtype=object)
    detected = parse_klapper_layout(raw, jenjang)
    if detected is not None:
        return detected

    uploaded_file.seek(0)
    simple = pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=object)
    return normalize_dataframe(simple, jenjang)


def parse_klapper_layout(raw: pd.DataFrame, jenjang: str) -> Optional[pd.DataFrame]:
    if raw is None or raw.empty:
        return None

    # Cari baris subheader yang berisi Urut dan Induk.
    subheader_row = None
    search_rows = min(15, len(raw))
    for r in range(search_rows):
        values = [clean_text(v).lower() for v in raw.iloc[r].tolist()]
        if "urut" in values and "induk" in values:
            subheader_row = r
            break

    if subheader_row is None:
        return None

    n_class = jumlah_kelas(jenjang)
    columns = all_columns(jenjang)
    start = subheader_row + 1
    rows = []
    for _, row in raw.iloc[start:].iterrows():
        # Posisi mengikuti format screenshot/template:
        # A No Urut, B No Induk, C NISN, D Nama siswa, E L/P, F TTL, G Ortu,
        # H dst kelas, lalu Tanggal Keluar Sekolah, Catatan.
        values = row.tolist()
        record = {}
        for idx, col in enumerate(columns):
            record[col] = values[idx] if idx < len(values) else ""
        rows.append(record)

    df = pd.DataFrame(rows, columns=columns)
    for col in df.columns:
        df[col] = df[col].apply(format_excel_date)

    return remove_empty_rows(df)


# =========================
# Tampilan preview HTML
# =========================

def dataframe_to_display_rows(df: pd.DataFrame, jenjang: str, min_rows: int = 10) -> list[list[str]]:
    cols = all_columns(jenjang)
    records = []
    for _, row in df[cols].iterrows():
        records.append([format_excel_date(row.get(col, "")) for col in cols])

    while len(records) < min_rows:
        records.append(["" for _ in cols])
    return records


def render_klapper_html(df: pd.DataFrame, jenjang: str, judul_sekolah: str, max_preview_rows: int = 12) -> str:
    n_class = jumlah_kelas(jenjang)
    class_numbers = "".join(f"<th>{i}</th>" for i in range(1, n_class + 1))
    rows = dataframe_to_display_rows(df.head(max_preview_rows), jenjang, min_rows=min(max_preview_rows, 12))

    body_rows = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(clean_text(v))}</td>" for v in row)
        body_rows.append(f"<tr>{cells}</tr>")

    return f"""
    <style>
    .klapper-wrapper {{
        overflow-x: auto;
        border: 1px solid #ddd;
        padding: 10px;
        background: white;
    }}
    table.klapper {{
        border-collapse: collapse;
        min-width: 1100px;
        width: 100%;
        font-family: 'Times New Roman', serif;
        color: #000;
    }}
    table.klapper th, table.klapper td {{
        border: 1px solid #000;
        padding: 5px 6px;
        text-align: center;
        vertical-align: middle;
        height: 26px;
        font-size: 14px;
    }}
    table.klapper .title td {{
        border: none;
        font-weight: bold;
        font-size: 20px;
        height: 26px;
    }}
    table.klapper .school td {{
        border: none;
        font-weight: bold;
        font-size: 20px;
        height: 26px;
    }}
    table.klapper .blank-title td {{
        border: none;
        height: 16px;
    }}
    table.klapper th {{
        font-weight: normal;
        background: #fff;
    }}
    table.klapper td:nth-child(4), table.klapper td:nth-child(6), table.klapper td:nth-child(7), table.klapper td:last-child {{
        text-align: left;
    }}
    </style>
    <div class="klapper-wrapper">
    <table class="klapper">
        <tr class="title"><td colspan="{len(all_columns(jenjang))}">BUKU KLAPPER SISWA</td></tr>
        <tr class="school"><td colspan="{len(all_columns(jenjang))}">{html.escape(judul_sekolah.upper())}</td></tr>
        <tr class="blank-title"><td colspan="{len(all_columns(jenjang))}"></td></tr>
        <tr>
            <th colspan="3">Nomor</th>
            <th rowspan="2">Nama siswa</th>
            <th rowspan="2">L/P</th>
            <th rowspan="2">Tempat Tanggal Lahir</th>
            <th rowspan="2">Nama Orang Tua</th>
            <th colspan="{n_class}">Tanggal Naik/Masuk Kelas</th>
            <th>Tanggal Keluar</th>
            <th rowspan="2">Catatan</th>
        </tr>
        <tr>
            <th>Urut</th>
            <th>Induk</th>
            <th>NISN</th>
            {class_numbers}
            <th>Sekolah</th>
        </tr>
        {''.join(body_rows)}
    </table>
    </div>
    """


# =========================
# Export Excel format cetak
# =========================

def export_klapper_excel(df: pd.DataFrame, jenjang: str, judul_sekolah: str, min_print_rows: int = 24) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
    
    n_class = jumlah_kelas(jenjang)
    cols = all_columns(jenjang)
    n_cols = len(cols)
    last_col_letter = get_column_letter(n_cols)
    class_start_col = 8  # H
    class_end_col = class_start_col + n_class - 1
    keluar_col = class_end_col + 1
    catatan_col = class_end_col + 2

    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Klapper"

    # Ukuran kolom mengikuti contoh template pengguna.
    widths = {
        1: 13,  # Urut
        2: 13,  # Induk
        3: 13,  # NISN
        4: 23,  # Nama siswa
        5: 5.7,  # L/P
        6: 19.5,  # Tempat Tanggal Lahir
        7: 18.5,  # Nama Orang Tua
        keluar_col: 15.6,
        catatan_col: 11.5,
    }
    for c in range(class_start_col, class_end_col + 1):
        widths[c] = 6
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 12)

    # Tinggi baris agar serupa template.
    ws.row_dimensions[1].height = 21
    ws.row_dimensions[2].height = 21
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24

    # Judul.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(1, 1, "BUKU KLAPPER SISWA")
    ws.cell(2, 1, judul_sekolah.upper())
    title_font = Font(name="Times New Roman", bold=True, size=14)
    header_font = Font(name="Times New Roman", size=11)
    body_font = Font(name="Times New Roman", size=11)
    for r in [1, 2]:
        ws.cell(r, 1).font = title_font
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

    # Header bertingkat.
    ws.merge_cells("A4:C4")
    ws.merge_cells("D4:D5")
    ws.merge_cells("E4:E5")
    ws.merge_cells("F4:F5")
    ws.merge_cells("G4:G5")
    ws.merge_cells(start_row=4, start_column=class_start_col, end_row=4, end_column=class_end_col)
    ws.merge_cells(start_row=4, start_column=catatan_col, end_row=5, end_column=catatan_col)

    ws.cell(4, 1, "Nomor")
    ws.cell(5, 1, "Urut")
    ws.cell(5, 2, "Induk")
    ws.cell(5, 3, "NISN")
    ws.cell(4, 4, "Nama siswa")
    ws.cell(4, 5, "L/P")
    ws.cell(4, 6, "Tempat Tanggal Lahir")
    ws.cell(4, 7, "Nama Orang Tua")
    ws.cell(4, class_start_col, "Tanggal Naik/Masuk Kelas")

    for i in range(1, n_class + 1):
        ws.cell(5, class_start_col + i - 1, i)

    ws.cell(4, keluar_col, "Tanggal Keluar")
    ws.cell(5, keluar_col, "Sekolah")
    ws.cell(4, catatan_col, "Catatan")

    # Style header dan tabel.
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = border
            cell.alignment = center
            cell.font = header_font

    clean_df = normalize_dataframe(df, jenjang)
    clean_df = remove_empty_rows(clean_df)
    records = dataframe_to_display_rows(clean_df, jenjang, min_rows=max(min_print_rows, len(clean_df)))

    start_row = 6
    for r_idx, record in enumerate(records, start=start_row):
        ws.row_dimensions[r_idx].height = 22
        for c_idx, value in enumerate(record, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.font = body_font
            if c_idx in [4, 6, 7, catatan_col]:
                cell.alignment = left
            else:
                cell.alignment = center

    # Freeze dan print setup.
    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:5"

# =========================
# Watermark / identitas pembuat + LinkedIn
# =========================
    last_row = start_row + len(records) - 1

    watermark_row = last_row + 2
    linkedin_row = last_row + 3

    linkedin_url = "www.linkedin.com/in/tun-pasek"

# Baris nama pembuat
    ws.merge_cells(
        start_row=watermark_row,
        start_column=1,
        end_row=watermark_row,
        end_column=n_cols
    )

    ws.cell(
        watermark_row,
        1,
        "Tun Pasek Sarwiko Dipranoto - Universitas Udayana"
    )

    watermark_cell = ws.cell(watermark_row, 1)
    watermark_cell.font = Font(
        name="Times New Roman",
        size=9,
        italic=True,
        color="808080"
    )
    watermark_cell.alignment = Alignment(
        horizontal="right",
        vertical="center"
    )

# Baris link LinkedIn
    ws.merge_cells(
        start_row=linkedin_row,
        start_column=1,
        end_row=linkedin_row,
        end_column=n_cols
    )

    ws.cell(
        linkedin_row,
        1,
        f"LinkedIn: {linkedin_url}"
    )

    linkedin_cell = ws.cell(linkedin_row, 1)
    linkedin_cell.hyperlink = linkedin_url
    linkedin_cell.font = Font(
        name="Times New Roman",
        size=9,
        italic=True,
        underline="single",
        color="0563C1"
    )
    linkedin_cell.alignment = Alignment(
        horizontal="right",
        vertical="center"
    )

    ws.row_dimensions[watermark_row].height = 18
    ws.row_dimensions[linkedin_row].height = 18

# Area print ikut memasukkan watermark dan link LinkedIn
    ws.print_area = f"A1:{last_col_letter}{linkedin_row}"

    # Sheet bantuan upload data mentah.
    guide = wb.create_sheet("Panduan Upload")
    guide["A1"] = "Format upload yang bisa digunakan"
    guide["A1"].font = Font(name="Calibri", bold=True, size=13)
    guide["A2"] = "1. Bisa upload format Buku Klapper seperti sheet ini, dengan header Nomor/Urut/Induk/NISN."
    guide["A3"] = "2. Bisa juga upload data biasa dengan baris pertama berisi nama kolom."
    guide["A4"] = "3. Untuk SD kolom kelas 1-6, untuk SMP/SMA kolom kelas 1-3."
    guide["A5"] = "4. Data dapat diurutkan otomatis berdasarkan No Urut, No Induk, NISN, atau Nama siswa."
    guide["A7"] = "Kolom minimal data biasa:"
    for i, col in enumerate(cols, start=1):
        guide.cell(8, i, col)
        guide.column_dimensions[get_column_letter(i)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# =========================
# Main Streamlit App
# =========================

def build_column_config(jenjang: str) -> dict:
    config = {
        "No Urut": st.column_config.NumberColumn("No Urut", step=1, format="%d", width="small"),
        "No Induk": st.column_config.TextColumn("No Induk", width="medium"),
        "NISN": st.column_config.TextColumn("NISN", width="medium"),
        "Nama siswa": st.column_config.TextColumn("Nama siswa", width="large"),
        "L/P": st.column_config.SelectboxColumn("L/P", options=["L", "P"], width="small"),
        "Tempat Tanggal Lahir": st.column_config.TextColumn("Tempat Tanggal Lahir", width="large"),
        "Nama Orang Tua": st.column_config.TextColumn("Nama Orang Tua", width="large"),
        "Tanggal Keluar Sekolah": st.column_config.TextColumn("Tanggal Keluar Sekolah", width="medium"),
        "Catatan": st.column_config.TextColumn("Catatan", width="large"),
    }
    for col in class_columns(jenjang):
        config[col] = st.column_config.TextColumn(col, width="small")
    return config


def main() -> None:
    st.title("📘 Aplikasi Buku Klapper Data Dapodik Siswa")
    st.caption("Format export dibuat seperti tabel Buku Klapper pada Excel: judul, header bertingkat, merge cell, border, ukuran kolom, dan ukuran baris tetap.")

    with st.sidebar:
        st.header("Pengaturan")
        jenjang = st.selectbox("Pilih jenjang", ["SD", "SMP", "SMA"], index=0)
        default_title = "SD NEGERI ....." if jenjang == "SD" else f"{jenjang} NEGERI ....."
        judul_sekolah = st.text_input("Judul sekolah pada Excel", value=default_title)
        sort_by = st.selectbox("Urutkan berdasarkan", ["No Urut", "No Induk", "NISN", "Nama siswa"], index=0)
        reset_no = st.checkbox("Nomori ulang No Urut setelah sorting", value=True)
        min_rows = st.number_input("Minimal baris kosong/cetak di Excel", min_value=10, max_value=200, value=24, step=1)
        st.info(f"{jenjang} memakai kolom kelas 1-{jumlah_kelas(jenjang)}.")

    tab_input, tab_preview, tab_template = st.tabs(["Input Data", "Preview Format", "Download Template"])

    with tab_input:
        st.subheader("1. Upload Excel atau isi manual")
        uploaded = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

        source_df = None
        if uploaded is not None:
            try:
                xls = pd.ExcelFile(uploaded)
                sheet = st.selectbox("Pilih sheet", xls.sheet_names)
                uploaded.seek(0)
                source_df = read_excel_upload(uploaded, sheet, jenjang)
                st.success("File berhasil dibaca. Format Buku Klapper dan format data biasa sama-sama bisa dipakai.")
            except Exception as exc:
                st.error(f"File belum bisa dibaca: {exc}")

        if source_df is None:
            source_df = blank_dataframe(jenjang, 12)

        source_df = normalize_dataframe(source_df, jenjang)

        edited_df = st.data_editor(
            source_df,
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config=build_column_config(jenjang),
            key=f"editor_{jenjang}",
        )

        st.session_state["df"] = edited_df

    with tab_preview:
        st.subheader("2. Preview dan download")
        df = st.session_state.get("df", blank_dataframe(jenjang, 12))
        df = normalize_dataframe(df, jenjang)
        df = remove_empty_rows(df)
        df = sort_dataframe(df, sort_by)
        if reset_no and len(df) > 0:
            df["No Urut"] = range(1, len(df) + 1)

        st.write(f"Jumlah data: **{len(df)}** | Jenjang: **{jenjang}** | Pengurutan: **{sort_by}**")
        st.markdown(render_klapper_html(df, jenjang, judul_sekolah, max_preview_rows=12), unsafe_allow_html=True)
        st.caption("Preview di aplikasi menampilkan sebagian baris. File Excel hasil download tetap berisi semua data.")

        excel_bytes = export_klapper_excel(df, jenjang, judul_sekolah, min_print_rows=int(min_rows))
        safe_school = re.sub(r"[^a-zA-Z0-9]+", "_", judul_sekolah.lower()).strip("_") or "sekolah"
        filename = f"buku_klapper_{jenjang.lower()}_{safe_school}.xlsx"
        st.download_button(
            "⬇️ Download Excel Format Buku Klapper",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()
        st.subheader("Data hasil sorting")
        st.dataframe(df, use_container_width=True, hide_index=True)

    with tab_template:
        st.subheader("Template sesuai format cetak")
        st.write("Download template kosong, lalu isi datanya di Excel atau gunakan input manual di aplikasi.")
        template_df = blank_dataframe(jenjang, 0)
        template_bytes = export_klapper_excel(template_df, jenjang, judul_sekolah, min_print_rows=int(min_rows))
        st.download_button(
            f"⬇️ Download Template {jenjang}",
            data=template_bytes,
            file_name=f"template_buku_klapper_{jenjang.lower()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.write("Format kolom input manual:")
        st.dataframe(pd.DataFrame(columns=all_columns(jenjang)), hide_index=True, use_container_width=True)


if __name__ == "__main__":
    main()
